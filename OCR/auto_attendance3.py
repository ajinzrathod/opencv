# Jai Swaminarayan

# PRE REQUESITES
# An excel file with roll no inserted in it and sheet no as "sheet1"

from sys import exit
from re import compile, findall
import os.path
from openpyxl import load_workbook, styles
from datetime import datetime
from pytesseract import image_to_string
from tkinter import Tk, filedialog
import pygame
from PIL import Image


global selectedImages
selectedImages = []

global max_row
global max_col
global mySheetName


pygame.init()


def displayImage(screen, px, topleft, prior):
    # ensure that the rect always has positive width, height
    x, y = topleft
    width = pygame.mouse.get_pos()[0] - topleft[0]
    height = pygame.mouse.get_pos()[1] - topleft[1]
    if width < 0:
        x += width
        width = abs(width)
    if height < 0:
        y += height
        height = abs(height)

    # eliminate redundant drawing cycles (when mouse isn't moving)
    current = x, y, width, height
    if not (width and height):
        return current
    if current == prior:
        return current

    # draw transparent box and blit it onto canvas
    screen.blit(px, px.get_rect())
    im = pygame.Surface((width, height))
    im.fill((128, 128, 128))
    pygame.draw.rect(im, (32, 32, 32), im.get_rect(), 1)
    im.set_alpha(128)
    screen.blit(im, (x, y))
    pygame.display.flip()

    # return current box extents
    return (x, y, width, height)


def setup(path):
    px = pygame.image.load(path)
    screen = pygame.display.set_mode(px.get_rect()[2:])
    screen.blit(px, px.get_rect())
    pygame.display.flip()
    return screen, px


def mainLoop(screen, px):
    topleft = bottomright = prior = None
    n = 0
    while n != 1:
        for event in pygame.event.get():
            if event.type == pygame.MOUSEBUTTONUP:
                if not topleft:
                    topleft = event.pos
                else:
                    bottomright = event.pos
                    n = 1
        if topleft:
            prior = displayImage(screen, px, topleft, prior)
    return (topleft + bottomright)


def cropImage(img, imgCount):
    input_loc = img
    output_loc = "out" + str(imgCount) + ".png"
    screen, px = setup(input_loc)
    left, upper, right, lower = mainLoop(screen, px)

    # ensure output rect always has positive width, height
    if right < left:
        left, right = right, left
    if lower < upper:
        lower, upper = upper, lower
    im = Image.open(input_loc)
    im = im.crop((left, upper, right, lower))
    pygame.display.quit()
    im.save(output_loc)

    return output_loc


def browseExcelFile():
    root.update()
    filez = filedialog.askopenfilenames(parent=root,
                                        title='Choose a file',
                                        filetypes=(("XLSX files", "*.xlsx*"),
                                                   ("XLSX files", "*.xlsx*")))
    return filez


def browseFiles():
    global selectedImages
    root.update()
    filez = filedialog.askopenfilenames(parent=root,
                                        title='Choose a file',
                                        filetypes=(("Png files", "*.png*"),
                                                   ("Jpg files", "*.jpg*"),
                                                   ("Jpeg files", "*.jpeg*")))
    selectedImages = root.tk.splitlist(filez)
    print(selectedImages)

    # Reading text from image and appending to file
    imageToFile()


def imageToFile():
    global selectedImages

    finalResult = []

    # for img in images:
    imgCount = 1
    for img in selectedImages:
        # cropImage
        img = cropImage(img, imgCount)
        imgCount += 1

        # Reading image using cv2
        # img_cv = imread(img)

        # # Converting To RGB Format
        # img_rgb = cvtColor(img_cv, COLOR_BGR2RGB)
        img_rgb = Image.open(img)

        # Extracting text from image
        text = image_to_string(img_rgb)

        # splitting attendes to new line
        result = text.split("\n")

        # Appending result to finall result
        finalResult = result + finalResult

    # Opening file
    f = open("list.txt", 'w')

    # Writing names to file
    for x in finalResult:
        print(x)
        f.write(x)
        f.write("\n")

    # Closing file
    f.close()
    checkExcelFileAndInsert()


def checkExcelFileAndInsert():
    global max_row
    global max_col
    global mySheetName

    mySheetName = 'sheet1'

    # If File does'nt exists
    if not(os.path.exists(filename)):
        print(f"File '{filename}' Does NOT Exists")
        exit(1)

    # open an Excel file and return a workbook
    workbook = load_workbook(filename)

    # By default: Sheet is not found
    sheetFound = False

    # Checking if specified sheet exists in workbook
    for sheet in workbook.sheetnames:
        if sheet.lower() == mySheetName:
            mySheetName = sheet
            sheetFound = True
            break

    # if sheetname not found
    if not sheetFound:
        print("sheet1 not found")
        exit(0)
        # exit(0)

    # Creating object of the sheet
    mySheet = workbook[mySheetName]

    # Finding max record and columns in .xls file
    max_row = mySheet.max_row
    max_col = mySheet.max_column

    # Trying to find roll no automatically
    rollNoRow, rollNoCol = findRollNoCol(mySheet)

    # If roll no is not found automatically
    if rollNoCol is None:
        print("Cell not found !")
        while True:
            rollNoCell = input("Enter cell no of roll no"
                               "(Example: A4, B5)\n")
            rollNoRow, rollNoCol = rollNoCellIsValid(rollNoCell)
            if rollNoRow is not None:
                break

    # If roll no is found automatically
    else:
        print("Roll no found at [", rollNoRow, chr(rollNoCol + 64), "]")

    # Inserting attendance
    insertData(rollNoRow, rollNoCol, mySheet)

    # Saving workbook
    workbook.save(filename)

    # Closing workbook
    workbook.close()


def rollNoCellIsValid(rollNoCell):
    if(len(rollNoCell) == 2):
        rollNoCell = rollNoCell.upper()

        # Initializing
        checkAlphabets = True
        checkNumbers = False
        cellValid = False

        # First Index must be a letter
        if not(rollNoCell[0] >= 'A') and rollNoCell[0] <= 'Z':
            return None, None

        # Iterating entire string
        for i in range(1, len(rollNoCell)):
            if checkAlphabets and rollNoCell[i] >= 'A' and rollNoCell[i] <= 'Z':
                continue
            else:
                checkNumbers = True
                checkAlphabets = False

            # Checking Numbers at end
            if checkNumbers is True:
                if rollNoCell[i] >= '0' and rollNoCell[i] <= '9':
                    cellValid = True
                else:
                    cellValid = False

        # If Cell is true, split Char And Numbers
        # Ex: "A4" to 'A' and '4'
        if cellValid is True:
            temp = compile("([a-zA-Z]+)([0-9]+)")
            res = temp.match(rollNoCell).groups()

            # printing result
            col = ord(res[0]) - 64  # Ascii code of char
            row = int(res[1])

            return row, col

    # If length is not 2
    else:
        return None, None


def findRollNoCol(mySheet):

    # Possible Text in files
    rollList = ['roll no', 'roll n', 'roll no.', 'roll',
                'rollno', 'rolln', 'rollno.',
                'roll number', 'roll num.', 'roll num',
                'rollnumber', 'rollnum.', 'rollnum']

    # Determining Max Row Search of Possible texts
    if max_row > 25:
        fetchTill = 25
    else:
        fetchTill = max_row

    print("Fetching Roll No in first ", fetchTill, " rows...", end=" ")

    # Iterating till rows Determined
    for i in range(1, fetchTill + 1):

        # Iterating over each column
        for j in range(1, max_col + 1):
            cell = mySheet.cell(row=i, column=j)

            # Reading value from object
            cell_value = cell.value
            if not(isinstance(cell_value, str)):
                continue

            # Converting to lower case
            cell_value = cell_value.lower()

            # Searching Possible texts as substring in all cols
            for item in rollList:
                if item in cell_value:
                    print("Done")
                    # Returning row and col where "roll no" was found
                    return i, j

    # if rollNoCell not found
    print("Done")
    return None, None


def insertData(rollNoRow, rollNoCol, mySheet):
    # Getting current date
    date = datetime.now().strftime('%d')
    month = datetime.now().strftime('%b')
    year = datetime.now().strftime('%Y')

    # Adding full stop so that when last digit is 0,
    # it doesnt consider last session as 0
    full_date = month + "\n" + date + ",\n" + year + "."

    # # Reading .txt File
    txtfile = "list.txt"
    if not(os.path.exists(txtfile)):
        print(f"File '{txtfile}' Does NOT Exists")
        exit(1)
    file = open("list.txt", "r")

    # rollNoRow Storing to another variable
    tempRollNoRow = rollNoRow

    # Default row, col for "list.txt"
    row, col = 0, 0

    # By Default, considering that only one lecture was held
    multipleSessions = False

    # Getting date of last attendance recorded
    lastDateAt = mySheet.cell(row=rollNoRow, column=max_col)
    lastDateAt.alignment = styles.Alignment(wrapText=True)

    # Checking if attendance of same date exists
    if (lastDateAt.value is not None) and (full_date in lastDateAt.value):
        print("\nNOTE: More than 1 sessions found")
        print("So, naming column name as S-1, S-2, ..\n")

        # Multiple lectures were held
        multipleSessions = True

        # Getting last digit of last session
        if lastDateAt.value[-1] >= '0' and lastDateAt.value[-1] <= '9':
            lastSessionNo = int(lastDateAt.value[-1])
        else:
            # Renaming cellvalue to session1
            lastSessionNo = 1
            lastDateAt.value = full_date + "\n S-1"

    # Insert date at last column
    insertDateAt = mySheet.cell(row=rollNoRow, column=max_col+1)
    insertDateAt.alignment = styles.Alignment(wrapText=True,
                                              horizontal='center')

    # If Multiple lectures were held
    if multipleSessions:
        insertDateAt.value = full_date + "\n S-" + str(lastSessionNo + 1)

    # if Multiple lectures were not held
    else:
        insertDateAt.value = full_date

    # Iterating over each attendee
    for attendee in file:

        # Remove blank lines
        if not attendee.strip():
            continue

        # Finding Roll no from entire String
        roll = findall("\d+", attendee)

        # Teachers wont have roll no, so ignoring them
        if not len(roll):
            print("Skipping", attendee, end="")
            continue

        # rollNoRow must be changed after each Iteration
        # so giving back the same value
        rollNoRow = tempRollNoRow

        # Iterating over all roll no of class
        for i in range(rollNoRow + 1, max_row + 1):

            # Getting cell_object of roll no of classes
            cell = mySheet.cell(row=i, column=rollNoCol)

            # NOTE: roll[0] is a `string`
            # If roll no of text file matches the roll no of xls file
            if(int(roll[0]) == cell.value):
                print("Inserteddddd")

                # Getting cell_object where new attendance will be recorded
                insertAtCell = mySheet.cell(row=i, column=max_col+1)

                # Store only first Number found from list
                insertAtCell.value = 'P'
                insertAtCell.alignment = styles.Alignment(
                    horizontal='center')

                # Printing those rollno who were recorded as present
                print(roll[0])
                break

            # Going to next record in .xls file
            rollNoRow += 1

        # Going to next record in .txt file
        row += 1


if __name__ == '__main__':

    # Creating root frame
    root = Tk()

    # Creating title
    root.title("Shri Hari")

    # Restricting to resize window in both X and Y direction
    root.resizable(False, False)

    # Set window background color
    root.config(background="#1ea")

    filename = browseExcelFile()
    filename = filename[0]

    browseFiles()

    root.mainloop()
