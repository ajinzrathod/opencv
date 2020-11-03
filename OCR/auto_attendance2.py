# Jai Swaminarayan
from sys import exit
import re
import os
import openpyxl
from datetime import datetime
from cv2 import imread, cvtColor, COLOR_BGR2RGB
import pytesseract
from tkinter import Tk, Button, filedialog
import pygame
from PIL import Image
global selectedImages
selectedImages = []
global max_row
global max_col
global mySheetName
pygame.init()
def displayImage(screen, px, topleft, prior):
    x, y = topleft
    width = pygame.mouse.get_pos()[0] - topleft[0]
    height = pygame.mouse.get_pos()[1] - topleft[1]
    if width < 0:
        x += width
        width = abs(width)
    if height < 0:
        y += height
        height = abs(height)
    current = x, y, width, height
    if not (width and height):
        return current
    if current == prior:
        return current
    screen.blit(px, px.get_rect())
    im = pygame.Surface((width, height))
    im.fill((128, 128, 128))
    pygame.draw.rect(im, (32, 32, 32), im.get_rect(), 1)
    im.set_alpha(128)
    screen.blit(im, (x, y))
    pygame.display.flip()
    return (x, y, width, height)
def setup(path):
    px = pygame.image.load(path)
    screen = pygame.display.set_mode(px.get_rect()[2:])
    screen.blit(px, px.get_rect())
    pygame.display.flip()
    return screen, px
def mainLoop(screen, px):
    topleft = bottomright = prior = None
    n = 0
    while n != 1:
        for event in pygame.event.get():
            if event.type == pygame.MOUSEBUTTONUP:
                if not topleft:
                    topleft = event.pos
                else:
                    bottomright = event.pos
                    n = 1
        if topleft:
            prior = displayImage(screen, px, topleft, prior)
    return (topleft + bottomright)
def cropImage(img, imgCount):
    input_loc = img
    output_loc = "out" + str(imgCount) + ".png"
    screen, px = setup(input_loc)
    left, upper, right, lower = mainLoop(screen, px)
    if right < left:
        left, right = right, left
    if lower < upper:
        lower, upper = upper, lower
    im = Image.open(input_loc)
    im = im.crop((left, upper, right, lower))
    pygame.display.quit()
    im.save(output_loc)
    return output_loc
def browseExcelFile():
    root.update()
    filez = filedialog.askopenfilenames(parent=root,
                                        title='Choose a file',
                                        filetypes=(("XLSX files", "*.xlsx*"),
                                                   ("XLSX files", "*.xlsx*")))
    return filez
def browseFiles():
    global selectedImages
    root.update()
    filez = filedialog.askopenfilenames(parent=root,
                                        title='Choose a file',
                                        filetypes=(("Png files", "*.png*"),
                                                   ("Jpg files", "*.jpg*"),
                                                   ("Jpeg files", "*.jpeg*")))
    selectedImages = root.tk.splitlist(filez)
    imageToFile()
def imageToFile():
    global selectedImages
    finalResult = []
    imgCount = 1
    for img in selectedImages:
        img = cropImage(img, imgCount)
        imgCount += 1
        img_cv = imread(img)
        img_rgb = cvtColor(img_cv, COLOR_BGR2RGB)
        text = pytesseract.image_to_string(img_rgb)
        result = text.split("\n")
        finalResult = result + finalResult
    f = open("list.txt", 'w')
    for x in finalResult:
        f.write(x)
        f.write("\n")
    f.close()
    checkExcelFileAndInsert()
def checkExcelFileAndInsert():
    global max_row
    global max_col
    global mySheetName
    mySheetName = 'sheet1'
    if not(os.path.exists(filename)):
        exit(1)
    workbook = openpyxl.load_workbook(filename)
    sheetFound = False
    for sheet in workbook.sheetnames:
        if sheet.lower() == mySheetName:
            mySheetName = sheet
            sheetFound = True
            break
    if not sheetFound:
        exit(0)
    mySheet = workbook[mySheetName]
    max_row = mySheet.max_row
    max_col = mySheet.max_column
    rollNoRow, rollNoCol = findRollNoCol(mySheet)
    insertData(rollNoRow, rollNoCol, mySheet)
    workbook.save(filename)
    workbook.close()


def findRollNoCol(mySheet):
    rollList = ['roll no', 'roll n', 'roll no.', 'roll',
                'rollno', 'rolln', 'rollno.',
                'roll number', 'roll num.', 'roll num',
                'rollnumber', 'rollnum.', 'rollnum']
    if max_row > 25:
        fetchTill = 25
    else:
        fetchTill = max_row
    for i in range(1, fetchTill + 1):
        for j in range(1, max_col + 1):
            cell = mySheet.cell(row=i, column=j)
            cell_value = cell.value
            if not(isinstance(cell_value, str)):
                continue
            cell_value = cell_value.lower()
            for item in rollList:
                if item in cell_value:
                    return i, j
    return None, None
def insertData(rollNoRow, rollNoCol, mySheet):
    date = datetime.now().strftime('%d')
    month = datetime.now().strftime('%b')
    year = datetime.now().strftime('%Y')
    full_date = month + "\n" + date + ",\n" + year + "."
    txtfile = "list.txt"
    if not(os.path.exists(txtfile)):
        exit(1)
    file = open("list.txt", "r")
    tempRollNoRow = rollNoRow
    row, col = 0, 0
    multipleSessions = False
    lastDateAt = mySheet.cell(row=rollNoRow, column=max_col)
    lastDateAt.alignment = openpyxl.styles.Alignment(wrapText=True)
    if (lastDateAt.value is not None) and (full_date in lastDateAt.value):
        multipleSessions = True
        if lastDateAt.value[-1] >= '0' and lastDateAt.value[-1] <= '9':
            lastSessionNo = int(lastDateAt.value[-1])
        else:
            lastSessionNo = 1
            lastDateAt.value = full_date + "\n S-1"
    insertDateAt = mySheet.cell(row=rollNoRow, column=max_col+1)
    insertDateAt.alignment = openpyxl.styles.Alignment(wrapText=True,
                                                       horizontal='center')
    if multipleSessions:
        insertDateAt.value = full_date + "\n S-" + str(lastSessionNo + 1)
    else:
        insertDateAt.value = full_date
    for attendee in file:
        if not attendee.strip():
            continue
        roll = re.findall("\d+", attendee)
        if not len(roll):
            continue
        rollNoRow = tempRollNoRow
        for i in range(rollNoRow + 1, max_row + 1):
            # Getting cell_object of roll no of classes
            cell = mySheet.cell(row=i, column=rollNoCol)
            if(int(roll[0]) == cell.value):
                insertAtCell = mySheet.cell(row=i, column=max_col+1)
                insertAtCell.value = 'P'
                insertAtCell.alignment = openpyxl.styles.Alignment(horizontal='center')
                break
            rollNoRow += 1
        row += 1
if __name__ == '__main__':
    root = Tk()
    root.title("Shri Hari")
    root.resizable(False, False)
    root.config(background="#1ea")
    button_exit = Button(root,
                         text="Exit",
                         command=exit)
    button_exit.grid(column=1, row=3)
    filename = browseExcelFile()
    filename = filename[0]
    browseFiles()
    root.mainloop()
